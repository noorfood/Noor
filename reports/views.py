from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from django.db.models import Sum, Avg, Count
from accounts.mixins import get_current_user, role_required
from production.models import MillingBatch, PackagingBatch
from clean_store.models import CleanRawIssuance, CleanRawReturn
from cleaning.models import CleanRawReceipt
from finished_store.models import FinishedGoodsReceipt, FinishedGoodsIssuance, FinishedGoodsReturn, PRODUCT_SIZE_CHOICES
from sales.models import SalesRecord
from reconciliation.models import MoneyReceipt, ReconciliationFlag
from accounts.models import User
from audit.models import AuditLog
from procurement.models import RawMaterialReceipt
import datetime
from django.db import transaction
from django.contrib import messages
from .models import MonthlySnapshot

@role_required('manager', 'md')
def dashboard(request):
    user = get_current_user(request)

    # Quick stats (Operational)
    total_batches = MillingBatch.objects.count() + PackagingBatch.objects.count()
    flagged_batches = MillingBatch.objects.filter(flag_level__in=['warning', 'critical']).count()
    open_flags = ReconciliationFlag.objects.filter(resolved=False).count()

    # Calculate global outstanding for Production (bags in hand not yet milled)
    accepted = CleanRawIssuance.objects.filter(status='accepted').aggregate(t=Sum('num_bags'))['t'] or 0
    milled_new = MillingBatch.objects.aggregate(t=Sum('bags_milled_new'))['t'] or 0
    milled_old = MillingBatch.objects.aggregate(t=Sum('outstanding_bags_milled'))['t'] or 0
    returned = CleanRawReturn.objects.filter(status__in=['pending', 'accepted']).aggregate(t=Sum('num_bags'))['t'] or 0
    outstanding_production = max(0, accepted - milled_new - milled_old - returned)

    # ─────────────────────────────────────────────────────────────────
    # OVERARCHING FINANCIAL METRICS (MD ONLY)
    # ─────────────────────────────────────────────────────────────────
    # ─────────────────────────────────────────────────────────────────
    # OVERARCHING FINANCIAL METRICS — split by Material Type (MD ONLY)
    # ─────────────────────────────────────────────────────────────────
    md_metrics = {
        'maize': {
            'produced': 0, 'in_store': 0, 'issued': 0, 'sold': 0,
            'money_received': 0.0, 'outstanding': 0.0, 'inventory_value': 0.0
        },
        'wheat': {
            'produced': 0, 'in_store': 0, 'issued': 0, 'sold': 0,
            'money_received': 0.0, 'outstanding': 0.0, 'inventory_value': 0.0
        }
    }

    from sales.models import SalesRecord, DirectSalePayment, SalesManagerCollection, SalesManagerPayment, SalesResult
    from production.models import BrandSale
    from finished_store.views import _fg_balance
    from finished_store.models import FinishedGoodsIssuance
    from pricing.models import PriceConfig

    if user.role == 'md':
        pass

    from django.utils import timezone
    import datetime

    f_from = request.GET.get('date_from', '')
    f_to   = request.GET.get('date_to', '')

    # Default to current month if NO filters at all
    now = timezone.now()
    if not f_from and not f_to:
        f_from = now.replace(day=1).date().isoformat()
        f_to   = now.date().isoformat()
    
    # Parse dates for query use
    date_from_obj = datetime.date.fromisoformat(f_from) if f_from else None
    date_to_obj   = datetime.date.fromisoformat(f_to) if f_to else None

    for mat in ['maize', 'wheat']:
        # 1. Produced (PackagingBatch) - Filtered
        batch_qs = PackagingBatch.objects.filter(material_type=mat)
        if date_from_obj: batch_qs = batch_qs.filter(date__gte=date_from_obj)
        if date_to_obj:   batch_qs = batch_qs.filter(date__lte=date_to_obj)
        md_metrics[mat]['produced'] = batch_qs.aggregate(t=Sum('qty_10kg'))['t'] or 0

        # 2. In Store - Always All-time (Current Balance)
        md_metrics[mat]['in_store'] = _fg_balance(mat, '10kg')

        # 3. Issued (Accepted FG Issuances) - Filtered
        iss_qs = FinishedGoodsIssuance.objects.filter(material_type=mat, status='accepted')
        if date_from_obj: iss_qs = iss_qs.filter(date__gte=date_from_obj)
        if date_to_obj:   iss_qs = iss_qs.filter(date__lte=date_to_obj)
        md_metrics[mat]['issued'] = iss_qs.aggregate(t=Sum('qty_issued'))['t'] or 0

        # 4. Sold - Filtered
        sm_res_qs = SalesResult.objects.filter(material_type=mat)
        if date_from_obj: sm_res_qs = sm_res_qs.filter(date__gte=date_from_obj)
        if date_to_obj:   sm_res_qs = sm_res_qs.filter(date__lte=date_to_obj)
        sm_sold_eq = float(sum(r.equivalent_sacks_sold for r in sm_res_qs))
        
        ds_sold_qs = DirectSalePayment.objects.filter(material_type=mat, status='confirmed')
        if date_from_obj: ds_sold_qs = ds_sold_qs.filter(date__gte=date_from_obj)
        if date_to_obj:   ds_sold_qs = ds_sold_qs.filter(date__lte=date_to_obj)
        ds_sold = ds_sold_qs.aggregate(t=Sum('qty_sold'))['t'] or 0
        md_metrics[mat]['sold'] = sm_sold_eq + float(ds_sold)

        # 5. Money Received & Outstanding
        # Money Received is period-based
        ds_qs_period = DirectSalePayment.objects.filter(material_type=mat, status='confirmed')
        if date_from_obj: ds_qs_period = ds_qs_period.filter(date__gte=date_from_obj)
        if date_to_obj:   ds_qs_period = ds_qs_period.filter(date__lte=date_to_obj)
        for ds in ds_qs_period:
            md_metrics[mat]['money_received'] += float(ds.total_received)

        # Outstanding is always All-time (unpaid debt)
        ds_qs_all = DirectSalePayment.objects.filter(material_type=mat, status='confirmed')
        for ds in ds_qs_all:
            md_metrics[mat]['outstanding'] += float(ds.outstanding)

        # Inventory Value
        sm_price = PriceConfig.get_active_price('sales_team', mat, '10kg', now.date()) or 0
        md_metrics[mat]['inventory_value'] = float(md_metrics[mat]['in_store'] * sm_price)

        # Calculate SM Money dynamically for Maize/Wheat
        for sm in User.objects.filter(role='sales_manager', status='active'):
            # ALL-TIME for ratios & outstanding
            m_coll_all = SalesManagerCollection.objects.filter(sales_manager=sm, material_type='maize', status='accepted').aggregate(t=Sum('total_value'))['t'] or 0
            w_coll_all = SalesManagerCollection.objects.filter(sales_manager=sm, material_type='wheat', status='accepted').aggregate(t=Sum('total_value'))['t'] or 0
            
            m_comm_all = SalesResult.objects.filter(recorded_by=sm, material_type='maize').aggregate(t=Sum('commission_amount'))['t'] or 0
            w_comm_all = SalesResult.objects.filter(recorded_by=sm, material_type='wheat').aggregate(t=Sum('commission_amount'))['t'] or 0
            
            m_net_owed_all = float(m_coll_all) - float(m_comm_all)
            w_net_owed_all = float(w_coll_all) - float(w_comm_all)
            total_net_all = m_net_owed_all + w_net_owed_all
            
            from sales.models import SalesManagerPayment
            total_paid_all = SalesManagerPayment.objects.filter(sales_manager=sm, status='confirmed').aggregate(t=Sum('amount_cash') + Sum('amount_transfer'))['t'] or 0
            total_paid_all = float(total_paid_all)
            
            if total_net_all > 0:
                m_ratio = m_net_owed_all / total_net_all
                w_ratio = w_net_owed_all / total_net_all
            else:
                m_ratio, w_ratio = 0.5, 0.5
                
            m_paid_all = total_paid_all * m_ratio
            w_paid_all = total_paid_all * w_ratio
            
            md_metrics['maize']['outstanding'] += max(0, m_net_owed_all - m_paid_all)
            md_metrics['wheat']['outstanding'] += max(0, w_net_owed_all - w_paid_all)
            
            # PERIOD-BASED for money_received
            sm_pay_qs = SalesManagerPayment.objects.filter(sales_manager=sm, status='confirmed')
            if date_from_obj: sm_pay_qs = sm_pay_qs.filter(date__gte=date_from_obj)
            if date_to_obj:   sm_pay_qs = sm_pay_qs.filter(date__lte=date_to_obj)
            
            total_paid_period = sm_pay_qs.aggregate(t=Sum('amount_cash') + Sum('amount_transfer'))['t'] or 0
            total_paid_period = float(total_paid_period)
            
            md_metrics['maize']['money_received'] += (total_paid_period * m_ratio)
            md_metrics['wheat']['money_received'] += (total_paid_period * w_ratio)

    staff_roster = User.objects.filter(status='active').exclude(role='md').order_by('role', 'full_name')
    audit_feed = AuditLog.objects.all().order_by('-timestamp')[:15]
    
    today_activities = AuditLog.objects.filter(
        user_id=user.pk,
        timestamp__date=datetime.date.today()
    ).order_by('-timestamp')

    recent_flagged = MillingBatch.objects.filter(flag_level__in=['warning', 'critical']).order_by('-date', '-created_at')[:5]

    # Production outstanding broken down by material
    from clean_store.models import CleanRawIssuance as _CRI, CleanRawReturn as _CRR
    def _prod_outstanding(mat):
        acc = _CRI.objects.filter(material_type=mat, status='accepted').aggregate(t=Sum('num_bags'))['t'] or 0
        mld_new = MillingBatch.objects.filter(material_type=mat).aggregate(t=Sum('bags_milled_new'))['t'] or 0
        mld_old = MillingBatch.objects.filter(material_type=mat).aggregate(t=Sum('outstanding_bags_milled'))['t'] or 0
        ret = _CRR.objects.filter(material_type=mat, status__in=['pending', 'accepted']).aggregate(t=Sum('num_bags'))['t'] or 0
        return max(0, acc - mld_new - mld_old - ret)
    outstanding_production_maize = _prod_outstanding('maize')
    outstanding_production_wheat = _prod_outstanding('wheat')

    # OM Retail and Bran Sales
    from sales.models import CompanyRetailLedger, SalesRecord
    from production.models import BrandSale
    from procurement.models import MATERIAL_CHOICES

    company_sales = SalesRecord.objects.filter(channel='company').order_by('-date', '-created_at')[:10]
    bran_sales = BrandSale.objects.all().order_by('-date', '-created_at')[:10]

    retail_balances = []
    for mat_val, mat_label in MATERIAL_CHOICES:
        pieces = CompanyRetailLedger.objects.filter(material_type=mat_val).aggregate(t=Sum('pieces_changed'))['t'] or 0
        retail_balances.append({'material': mat_label, 'pieces': pieces})

    # SM Accountability Summaries — expose maize/wheat separately
    from sales.views import get_sm_goods_holding, get_sm_money_outstanding
    from sales.models import SalesManagerCollection
    sm_list = User.objects.filter(role='sales_manager', status='active').order_by('full_name')
    sm_summaries = []
    for sm in sm_list:
        maize = get_sm_goods_holding(sm, 'maize')
        wheat = get_sm_goods_holding(sm, 'wheat')
        money = get_sm_money_outstanding(sm)
        pending_count = SalesManagerCollection.objects.filter(sales_manager=sm, status='pending').count()
        sm_summaries.append({
            'sm': sm,
            'maize_holding': maize,
            'wheat_holding': wheat,
            'goods_balance': maize + wheat,
            'money_outstanding': money,
            'pending_count': pending_count,
        })

    # Real-world handshake: Goods issued to Company channel but not yet GM-acknowledged
    from finished_store.models import FinishedGoodsIssuance
    pending_company_issuances = FinishedGoodsIssuance.objects.filter(channel='company', status='pending').order_by('-date', '-created_at')

    # Pending SM payments awaiting GM confirmation
    from sales.models import SalesManagerPayment
    pending_sm_payments = SalesManagerPayment.objects.filter(status='pending_gm').order_by('-date', '-created_at')

    # Pending Raw Material Cost Entries (MD Review)
    pending_raw_costs = RawMaterialReceipt.objects.filter(cost_status='pending') if user.role == 'md' else []

    return render(request, 'reports/dashboard.html', {
        'current_user': user,
        'total_batches': total_batches,
        'flagged_batches': flagged_batches,
        'open_flags': open_flags,
        'outstanding_production': outstanding_production,
        'outstanding_production_maize': outstanding_production_maize,
        'outstanding_production_wheat': outstanding_production_wheat,
        'md_metrics': md_metrics,
        'recent_flagged': recent_flagged,
        'staff_roster': staff_roster,
        'audit_feed': audit_feed,
        'company_sales': company_sales,
        'bran_sales': bran_sales,
        'retail_balances': retail_balances,
        'sm_summaries': sm_summaries,
        'pending_company_issuances': pending_company_issuances,
        'pending_sm_payments': pending_sm_payments,
        'pending_raw_costs': pending_raw_costs,
        'today_activities': today_activities,
    })


@role_required('manager', 'md')
def production_report(request):
    user = get_current_user(request)
    f_from = request.GET.get('date_from', '')
    f_to = request.GET.get('date_to', '')
    f_material = request.GET.get('material', '')
    f_flag = request.GET.get('flag', '')

    milling = MillingBatch.objects.all().order_by('-date', '-created_at')
    packaging = PackagingBatch.objects.all().order_by('-date', '-created_at')
    
    if f_from:
        milling = milling.filter(date__gte=f_from)
        packaging = packaging.filter(date__gte=f_from)
    if f_to:
        milling = milling.filter(date__lte=f_to)
        packaging = packaging.filter(date__lte=f_to)
    if f_material:
        milling = milling.filter(material_type=f_material)
        packaging = packaging.filter(material_type=f_material)
    if f_flag:
        milling = milling.filter(flag_level=f_flag)

    milling_totals = milling.aggregate(
        total_raw_kg=Sum('total_raw_kg'),
        total_powder_kg=Sum('bulk_powder_kg'),
        total_loss_kg=Sum('loss_kg'),
    )
    
    packaging_totals = packaging.aggregate(
        total_powder_used=Sum('powder_used_kg'),
        total_output_kg=Sum('total_output_kg'),
        total_loss_kg=Sum('loss_kg'),
    )

    if request.GET.get('export') == 'xlsx':
        return _export_production_xlsx(milling, packaging)

    return render(request, 'reports/production.html', {
        'current_user': user, 'milling': milling, 'packaging': packaging, 
        'milling_totals': milling_totals, 'packaging_totals': packaging_totals,
        'f_from': f_from, 'f_to': f_to, 'f_material': f_material, 'f_flag': f_flag,
    })


def _export_production_xlsx(milling, packaging):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    
    # Milling Sheet
    ws_m = wb.active
    ws_m.title = 'Milling Report'
    headers_m = ['Batch ID', 'Date', 'Officer', 'Material', 'Shift', 'Bags Collected', 'Bags Milled',
               'Outstanding', 'Powder KG', 'Loss KG', 'Loss %', 'Flag']
    ws_m.append(['NOOR FOODS - Milling Report'])
    ws_m.append(['Generated:', datetime.datetime.now().strftime('%Y-%m-%d %H:%M')])
    ws_m.append([])
    ws_m.append(headers_m)

    for b in milling:
        ws_m.append([
            b.pk, str(b.date), b.production_officer.full_name, b.material_type.upper(),
            b.shift, b.bags_milled_new, b.outstanding_bags_milled,
            float(b.bulk_powder_kg), float(b.loss_kg), float(b.loss_pct), b.flag_level.upper()
        ])
        
    # Packaging Sheet
    ws_p = wb.create_sheet('Packaging Report')
    headers_p = ['Batch ID', 'Date', 'Officer', 'Material', 'Shift', 'Milling Source ID', 'Powder Used KG', 'Sacks Produced (10kg)', 'Output KG', 'Loss KG', 'Loss %']
    ws_p.append(['NOOR FOODS - Packaging Report'])
    ws_p.append(['Generated:', datetime.datetime.now().strftime('%Y-%m-%d %H:%M')])
    ws_p.append([])
    ws_p.append(headers_p)
    
    for p in packaging:
        ws_p.append([
            p.pk, str(p.date), p.production_officer.full_name, p.material_type.upper(),
            p.shift, p.milling_batch_id, float(p.powder_used_kg), p.qty_10kg,
            float(p.total_output_kg), float(p.loss_kg), float(p.loss_pct)
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=production_report.xlsx'
    wb.save(response)
    return response


@role_required('manager', 'md')
def store_report(request):
    user = get_current_user(request)
    from clean_store.views import _get_clean_store_balance
    from finished_store.views import _fg_balance
    from procurement.models import RawMaterialReceipt, RawMaterialIssuance

    # --- Dirty Raw Store: bags received from market minus bags issued to cleaning ---
    materials = ['maize', 'wheat']
    dirty_raw = {}
    for mat in materials:
        received = RawMaterialReceipt.objects.filter(material_type=mat).aggregate(
            t=Sum('num_bags'))['t'] or 0
        issued = RawMaterialIssuance.objects.filter(material_type=mat).aggregate(
            t=Sum('num_bags_issued'))['t'] or 0
        dirty_raw[mat] = max(0, received - issued)

    # --- Clean Raw Store (post-cleaning, pre-production) ---
    clean_maize = _get_clean_store_balance('maize')
    clean_wheat = _get_clean_store_balance('wheat')

    # --- Finished Goods: balance per (material, size) ---
    sizes = [s for s, _ in PRODUCT_SIZE_CHOICES]
    fg_maize_rows = []
    fg_wheat_rows = []
    fg_totals_rows = []
    for size in sizes:
        # Query total accepted receipts directly by material_type
        maize_in = FinishedGoodsReceipt.objects.filter(
            product_size=size, material_type='maize', status='accepted'
        ).aggregate(t=Sum('qty_received'))['t'] or 0
        wheat_in = FinishedGoodsReceipt.objects.filter(
            product_size=size, material_type='wheat', status='accepted'
        ).aggregate(t=Sum('qty_received'))['t'] or 0
        
        # Query total accepted issuances by material_type
        maize_out = FinishedGoodsIssuance.objects.filter(
            product_size=size, material_type='maize', status='accepted'
        ).aggregate(t=Sum('qty_issued'))['t'] or 0
        wheat_out = FinishedGoodsIssuance.objects.filter(
            product_size=size, material_type='wheat', status='accepted'
        ).aggregate(t=Sum('qty_issued'))['t'] or 0
        
        # Calculate net balances (In - Out + Returns normally, but report simplifies to just In - Out)
        # Adding Returned back in for accuracy
        maize_ret = FinishedGoodsReturn.objects.filter(
            product_size=size, material_type='maize', status='accepted'
        ).aggregate(t=Sum('qty_returned'))['t'] or 0
        wheat_ret = FinishedGoodsReturn.objects.filter(
            product_size=size, material_type='wheat', status='accepted'
        ).aggregate(t=Sum('qty_returned'))['t'] or 0

        total_in = maize_in + wheat_in
        total_out = maize_out + wheat_out
        total_ret = maize_ret + wheat_ret
        net = max(0, total_in - total_out + total_ret)
        
        fg_maize_rows.append((size.upper(), maize_in))
        fg_wheat_rows.append((size.upper(), wheat_in))
        fg_totals_rows.append((size.upper(), net))


    # Zip into a single list of tuples for template iteration: (size, maize_in, wheat_in, net)
    fg_rows = [
        (m[0], m[1], w[1], t[1])
        for m, w, t in zip(fg_maize_rows, fg_wheat_rows, fg_totals_rows)
    ]

    return render(request, 'reports/store.html', {
        'current_user': user,
        'dirty_raw': dirty_raw,
        'clean_maize': clean_maize,
        'clean_wheat': clean_wheat,
        'fg_rows': fg_rows,
    })


@role_required('md')
def sales_report(request):
    user = get_current_user(request)
    f_from = request.GET.get('date_from', '')
    f_to = request.GET.get('date_to', '')
    f_person = request.GET.get('sales_person', '')
    f_material = request.GET.get('material', '')

    from sales.models import SalesPerson, SalesPayment, SalesResult, DirectSalePayment

    sales = SalesRecord.objects.all().order_by('-date', '-created_at')
    promoter_sales = SalesResult.objects.all().order_by('-date', '-created_at')
    direct_sales_confirmed = DirectSalePayment.objects.filter(status='confirmed').order_by('-date', '-created_at')

    if f_from:
        sales = sales.filter(date__gte=f_from)
        promoter_sales = promoter_sales.filter(date__gte=f_from)
        direct_sales_confirmed = direct_sales_confirmed.filter(date__gte=f_from)
    if f_to:
        sales = sales.filter(date__lte=f_to)
        promoter_sales = promoter_sales.filter(date__lte=f_to)
        direct_sales_confirmed = direct_sales_confirmed.filter(date__lte=f_to)
    if f_material:
        sales = sales.filter(material_type=f_material)
        promoter_sales = promoter_sales.filter(material_type=f_material)
        direct_sales_confirmed = direct_sales_confirmed.filter(material_type=f_material)

    # Overall totals
    total_value      = sum(float(r.total_value) for r in sales)
    total_commission = sum(float(r.commission_amount) for r in sales)
    total_net        = sum(float(r.net_payable) for r in sales)
    total_received   = sum(r.total_paid for r in sales)
    total_outstanding= sum(r.amount_outstanding for r in sales)

    # Per-material totals (Consolidated)
    maize_value = (float(sum(r.total_value for r in sales if r.material_type == 'maize')) +
                   float(sum(r.gross_value for r in promoter_sales if r.material_type == 'maize')) +
                   float(sum(r.total_sale_value for r in direct_sales_confirmed if r.material_type == 'maize')))
    
    wheat_value = (float(sum(r.total_value for r in sales if r.material_type == 'wheat')) +
                   float(sum(r.gross_value for r in promoter_sales if r.material_type == 'wheat')) +
                   float(sum(r.total_sale_value for r in direct_sales_confirmed if r.material_type == 'wheat')))
    
    maize_received = (float(sum(r.total_paid for r in sales if r.material_type == 'maize')) +
                      float(sum(r.amount_returned for r in promoter_sales if r.material_type == 'maize')) +
                      float(sum(r.total_received for r in direct_sales_confirmed if r.material_type == 'maize')))
    
    wheat_received = (float(sum(r.total_paid for r in sales if r.material_type == 'wheat')) +
                      float(sum(r.amount_returned for r in promoter_sales if r.material_type == 'wheat')) +
                      float(sum(r.total_received for r in direct_sales_confirmed if r.material_type == 'wheat')))

    # Commission offsets per material
    maize_comm = (float(sum(r.commission_amount for r in sales if r.material_type == 'maize')) +
                  float(sum(r.commission_amount for r in promoter_sales if r.material_type == 'maize')))
                  
    wheat_comm = (float(sum(r.commission_amount for r in sales if r.material_type == 'wheat')) +
                  float(sum(r.commission_amount for r in promoter_sales if r.material_type == 'wheat')))

    maize_outstanding = max(0, (maize_value - maize_comm) - maize_received)
    wheat_outstanding = max(0, (wheat_value - wheat_comm) - wheat_received)

    all_persons = SalesPerson.objects.all().order_by('name')

    v_record = float(sum(r.total_value for r in sales))
    v_promoted = float(sum(r.gross_value for r in promoter_sales))
    v_direct = float(sum(r.total_sale_value for r in direct_sales_confirmed))
    tv = v_record + v_promoted + v_direct

    c_record = float(sum(r.commission_amount for r in sales))
    c_promoted = float(sum(r.commission_amount for r in promoter_sales))
    tc = c_record + c_promoted
    
    r_record = float(sum(r.total_paid for r in sales))
    r_promoted = float(sum(r.amount_returned for r in promoter_sales))
    r_direct = float(sum(r.total_received for r in direct_sales_confirmed))
    tr = r_record + r_promoted + r_direct

    tn = tv - tc
    tout = max(0, tn - tr)

    return render(request, 'reports/sales.html', {
        'current_user': user,
        'sales': sales,
        'promoter_sales': promoter_sales,
        'direct_sales': direct_sales_confirmed,
        'total_value': tv,
        'total_commission': tc,
        'total_net': tn,
        'total_received': tr,
        'total_outstanding': tout,
        'maize_value': maize_value,
        'wheat_value': wheat_value,
        'maize_received': maize_received,
        'wheat_received': wheat_received,
        'maize_outstanding': maize_outstanding,
        'wheat_outstanding': wheat_outstanding,
        'all_persons': all_persons,
        'f_from': f_from, 'f_to': f_to, 'f_person': f_person, 'f_material': f_material,
    })


@role_required('manager', 'md')
def outstanding_report(request):
    user = get_current_user(request)

    # Production Balances: raw bags issued to a production officer but not yet milled
    from accounts.models import User
    from sales.views import get_sm_goods_holding, get_sm_money_outstanding
    from sales.models import SalesManagerCollection
    prod_users = User.objects.filter(role='production_officer', status='active')
    prod_balances = []
    for u in prod_users:
        # Per-material balance for production officer
        def _po_bal(mat):
            acc = CleanRawIssuance.objects.filter(issued_to=u, status='accepted', material_type=mat).aggregate(t=Sum('num_bags'))['t'] or 0
            mld = MillingBatch.objects.filter(production_officer=u, material_type=mat).aggregate(t=Sum('bags_milled_new'))['t'] or 0
            ret = CleanRawReturn.objects.filter(returned_by=u, material_type=mat, status__in=['pending', 'accepted']).aggregate(t=Sum('num_bags'))['t'] or 0
            return max(0, acc - mld - ret)
        maize_bal = _po_bal('maize')
        wheat_bal = _po_bal('wheat')
        total_bal = maize_bal + wheat_bal
        if total_bal > 0:
            prod_balances.append({'user': u, 'balance': total_bal, 'maize_balance': maize_bal, 'wheat_balance': wheat_bal})

    # Sales Balances: money outstanding per Sales Manager — with per-material goods split
    sales_managers = User.objects.filter(role='sales_manager', status='active').order_by('full_name')
    sales_balances = []
    for sm in sales_managers:
        maize_holding = get_sm_goods_holding(sm, 'maize')
        wheat_holding = get_sm_goods_holding(sm, 'wheat')
        goods_holding = maize_holding + wheat_holding
        money_out = get_sm_money_outstanding(sm)
        record_count = SalesManagerCollection.objects.filter(sales_manager=sm).count()

        if goods_holding > 0 or money_out > 0:
            sales_balances.append({
                'sm': sm,
                'maize_holding': maize_holding,
                'wheat_holding': wheat_holding,
                'goods_holding': goods_holding,
                'outstanding': money_out,
                'record_count': record_count,
            })

    recon_flags = ReconciliationFlag.objects.filter(resolved=False).order_by('-date', '-created_at')
    return render(request, 'reports/outstanding.html', {
        'current_user': user,
        'prod_outstanding': prod_balances,
        'sales_balances': sales_balances,
        'recon_flags': recon_flags,
    })


@role_required('manager', 'md')
def company_flow(request):
    """Full goods-to-money pipeline. Money totals only visible to MD."""
    user = get_current_user(request)
    show_money = (user.role == 'md')

    from clean_store.views import _get_clean_store_balance
    from finished_store.views import _fg_balance
    from procurement.models import RawMaterialReceipt, RawMaterialIssuance
    from accounts.models import User
    from sales.views import get_sm_goods_holding, get_sm_money_outstanding, get_gm_goods_holding, get_company_goods_holding
    from sales.models import (
        SalesManagerCollection, SalesManagerPayment, SalesRecord,
        CompanyRetailLedger, SalesResult, DirectSalePayment
    )
    from production.models import BrandSale

    # 2. Standardized Intelligence Window
    from django.utils import timezone
    from datetime import timedelta
    range_type = request.GET.get('range', 'month')
    now = timezone.now()
    
    start_date = None
    end_date = now.date()

    if range_type == 'week':
        start_date = (now - timedelta(days=now.weekday())).date()
    elif range_type == 'year':
        start_date = now.replace(month=1, day=1).date()
    elif range_type == 'all':
        start_date = None
        end_date = None
    else: # Default: month
        start_date = now.replace(day=1).date()

    # Manual Overrides
    f_from = request.GET.get('date_from', '')
    f_to = request.GET.get('date_to', '')
    if f_from: 
        try: start_date = datetime.date.fromisoformat(f_from)
        except: pass
    if f_to: 
        try: end_date = datetime.date.fromisoformat(f_to)
        except: pass

    # --- Query Helpers ---
    def _filter_date(qs, date_field='date'):
        if start_date:
            lookup = f"{date_field}__gte"
            qs = qs.filter(**{lookup: start_date})
        if end_date:
            lookup = f"{date_field}__lte"
            qs = qs.filter(**{lookup: end_date})
        return qs

    # Stage 1: Raw Store
    raw_maize_in  = _filter_date(RawMaterialReceipt.objects.filter(material_type='maize')).aggregate(t=Sum('num_bags'))['t'] or 0
    raw_wheat_in  = _filter_date(RawMaterialReceipt.objects.filter(material_type='wheat')).aggregate(t=Sum('num_bags'))['t'] or 0
    raw_maize_out = _filter_date(RawMaterialIssuance.objects.filter(material_type='maize')).aggregate(t=Sum('num_bags_issued'))['t'] or 0
    raw_wheat_out = _filter_date(RawMaterialIssuance.objects.filter(material_type='wheat')).aggregate(t=Sum('num_bags_issued'))['t'] or 0
    
    # Balance is ALWAYS ALL-TIME
    def _raw_bal(mat):
        i = RawMaterialReceipt.objects.filter(material_type=mat).aggregate(t=Sum('num_bags'))['t'] or 0
        o = RawMaterialIssuance.objects.filter(material_type=mat).aggregate(t=Sum('num_bags_issued'))['t'] or 0
        return max(0, i - o)
    raw_maize_bal = _raw_bal('maize')
    raw_wheat_bal = _raw_bal('wheat')

    # Stage 2: Clean Store
    clean_maize_in  = _filter_date(CleanRawReceipt.objects.filter(material_type='maize')).aggregate(t=Sum('num_bags'))['t'] or 0
    clean_wheat_in  = _filter_date(CleanRawReceipt.objects.filter(material_type='wheat')).aggregate(t=Sum('num_bags'))['t'] or 0
    clean_maize_bal = _get_clean_store_balance('maize')
    clean_wheat_bal = _get_clean_store_balance('wheat')
    clean_maize_out = _filter_date(CleanRawIssuance.objects.filter(material_type='maize')).aggregate(t=Sum('num_bags'))['t'] or 0
    clean_wheat_out = _filter_date(CleanRawIssuance.objects.filter(material_type='wheat')).aggregate(t=Sum('num_bags'))['t'] or 0

    # Stage 3: Production — split by material
    def _milled(mat):
        qs = MillingBatch.objects.filter(material_type=mat)
        qs = _filter_date(qs)
        n = qs.aggregate(t=Sum('bags_milled_new'))['t'] or 0
        o = qs.aggregate(t=Sum('outstanding_bags_milled'))['t'] or 0
        return n + o
    
    maize_bags_milled = _milled('maize')
    wheat_bags_milled = _milled('wheat')
    
    maize_powder_produced = float(_filter_date(MillingBatch.objects.filter(material_type='maize')).aggregate(t=Sum('bulk_powder_kg'))['t'] or 0)
    wheat_powder_produced = float(_filter_date(MillingBatch.objects.filter(material_type='wheat')).aggregate(t=Sum('bulk_powder_kg'))['t'] or 0)
    
    maize_powder_used = float(_filter_date(PackagingBatch.objects.filter(material_type='maize')).aggregate(t=Sum('powder_used_kg'))['t'] or 0)
    wheat_powder_used = float(_filter_date(PackagingBatch.objects.filter(material_type='wheat')).aggregate(t=Sum('powder_used_kg'))['t'] or 0)
    
    maize_sacks_produced  = _filter_date(PackagingBatch.objects.filter(material_type='maize')).aggregate(t=Sum('qty_10kg'))['t'] or 0
    wheat_sacks_produced  = _filter_date(PackagingBatch.objects.filter(material_type='wheat')).aggregate(t=Sum('qty_10kg'))['t'] or 0
    
    # Balance is ALWAYS ALL-TIME
    def _powder_bal(mat):
        p = float(MillingBatch.objects.filter(material_type=mat).aggregate(t=Sum('bulk_powder_kg'))['t'] or 0)
        u = float(PackagingBatch.objects.filter(material_type=mat).aggregate(t=Sum('powder_used_kg'))['t'] or 0)
        return max(0.0, p - u)
    maize_powder_bal = _powder_bal('maize')
    wheat_powder_bal = _powder_bal('wheat')

    prod_users = User.objects.filter(role='production_officer', status='active')
    prod_officer_rows = []
    for u in prod_users:
        acc_qs = CleanRawIssuance.objects.filter(issued_to=u, status='accepted')
        acc = _filter_date(acc_qs).aggregate(t=Sum('num_bags'))['t'] or 0
        
        mld_qs = MillingBatch.objects.filter(production_officer=u)
        mld = _filter_date(mld_qs).aggregate(a=Sum('bags_milled_new'), b=Sum('outstanding_bags_milled'))
        mld_total = (mld['a'] or 0) + (mld['b'] or 0)
        
        ret_qs = CleanRawReturn.objects.filter(returned_by=u, status__in=['pending', 'accepted'])
        ret = _filter_date(ret_qs).aggregate(t=Sum('num_bags'))['t'] or 0
        
        # Officer Balance is ALWAYS ALL-TIME
        acc_full = CleanRawIssuance.objects.filter(issued_to=u, status='accepted').aggregate(t=Sum('num_bags'))['t'] or 0
        mld_full = MillingBatch.objects.filter(production_officer=u).aggregate(a=Sum('bags_milled_new'), b=Sum('outstanding_bags_milled'))
        mld_f_total = (mld_full['a'] or 0) + (mld_full['b'] or 0)
        ret_full = CleanRawReturn.objects.filter(returned_by=u, status__in=['pending', 'accepted']).aggregate(t=Sum('num_bags'))['t'] or 0
        bal = max(0, acc_full - mld_f_total - ret_full)
        
        prod_officer_rows.append({'user': u, 'accepted': acc, 'milled': mld_total, 'returned': ret, 'balance': bal})

    # Stage 4: Finished Goods Store — split by material
    fg_maize_in  = _filter_date(FinishedGoodsReceipt.objects.filter(material_type='maize', status='accepted')).aggregate(t=Sum('qty_received'))['t'] or 0
    fg_wheat_in  = _filter_date(FinishedGoodsReceipt.objects.filter(material_type='wheat', status='accepted')).aggregate(t=Sum('qty_received'))['t'] or 0
    fg_maize_out = _filter_date(FinishedGoodsIssuance.objects.filter(material_type='maize', status='accepted')).aggregate(t=Sum('qty_issued'))['t'] or 0
    fg_wheat_out = _filter_date(FinishedGoodsIssuance.objects.filter(material_type='wheat', status='accepted')).aggregate(t=Sum('qty_issued'))['t'] or 0
    fg_maize_bal = _fg_balance('maize', '10kg')
    fg_wheat_bal = _fg_balance('wheat', '10kg')
    fg_in  = fg_maize_in + fg_wheat_in
    fg_out = fg_maize_out + fg_wheat_out
    fg_bal = fg_maize_bal + fg_wheat_bal

    # Stage 5: Brand Sales & Byproducts
    brand_sales = _filter_date(BrandSale.objects.all())
    total_brand_sacks = brand_sales.aggregate(t=Sum('qty_sacks'))['t'] or 0
    brand_sales_value = float(brand_sales.aggregate(t=Sum('total_amount'))['t'] or 0)
    brand_received    = float(brand_sales.aggregate(t=Sum('amount_cash'))['t'] or 0) + float(brand_sales.aggregate(t=Sum('amount_transfer'))['t'] or 0)
    brand_outstanding = max(0, brand_sales_value - brand_received)

    # Stage 6: Sales -> Money (Sales Manager Level) — with per-material split
    sales_managers = User.objects.filter(role='sales_manager', status='active').order_by('full_name')
    sm_rows = []
    total_sacks_collected   = 0
    total_goods_outstanding = 0
    total_money_outstanding = 0
    total_money_received    = 0
    total_sales_value       = 0

    for sm in sales_managers:
        # Goods collected per material (accepted)
        maize_coll = _filter_date(SalesManagerCollection.objects.filter(sales_manager=sm, status='accepted', material_type='maize'))
        wheat_coll = _filter_date(SalesManagerCollection.objects.filter(sales_manager=sm, status='accepted', material_type='wheat'))
        maize_collected = sum(c.qty_sacks for c in maize_coll)
        wheat_collected = sum(c.qty_sacks for c in wheat_coll)
        collected = maize_collected + wheat_collected
        
        coll_val = sum(c.total_value for c in maize_coll) + sum(c.total_value for c in wheat_coll)
        
        # Real Sold (Promoter level)
        res_qs = _filter_date(SalesResult.objects.filter(recorded_by=sm))
        sold_eq = float(sum(r.equivalent_sacks_sold for r in res_qs))
        res_val = res_qs.aggregate(t=Sum('gross_value'))['t'] or 0

        # Money received (confirmed)
        payments = _filter_date(SalesManagerPayment.objects.filter(sales_manager=sm, status='confirmed'))
        received = sum(p.total for p in payments)

        maize_bal = get_sm_goods_holding(sm, 'maize')
        wheat_bal = get_sm_goods_holding(sm, 'wheat')
        goods_bal = maize_bal + wheat_bal
        money_out = get_sm_money_outstanding(sm)

        total_sacks_collected   += collected
        total_goods_outstanding += goods_bal
        total_money_outstanding += money_out
        total_money_received    += received
        total_sales_value       += float(res_val) # Real revenue from results

        sm_rows.append({
            'sm': sm,
            'maize_collected': maize_collected,
            'wheat_collected': wheat_collected,
            'collected': collected,
            'sold_eq': sold_eq,
            'maize_balance': maize_bal,
            'wheat_balance': wheat_bal,
            'goods_balance': goods_bal,
            'money_outstanding': money_out,
            'money_received': received,
            'total_sales_value': float(res_val),
            'coll_value': coll_val,
        })

    # GM & MD Direct Sales Context (separate pools)
    gm_maize_hand = get_gm_goods_holding('maize')
    gm_wheat_hand = get_gm_goods_holding('wheat')
    
    # Per-user breakdown
    gm_users = User.objects.filter(role='manager', status='active')
    md_users = User.objects.filter(role='md', status='active')
    gm_user_maize = sum(get_company_goods_holding(u, 'maize') for u in gm_users)
    gm_user_wheat = sum(get_company_goods_holding(u, 'wheat') for u in gm_users)
    md_user_maize = sum(get_company_goods_holding(u, 'maize') for u in md_users)
    md_user_wheat = sum(get_company_goods_holding(u, 'wheat') for u in md_users)
    
    gm_maize_pieces = CompanyRetailLedger.objects.filter(material_type='maize').aggregate(t=Sum('pieces_changed'))['t'] or 0
    gm_wheat_pieces = CompanyRetailLedger.objects.filter(material_type='wheat').aggregate(t=Sum('pieces_changed'))['t'] or 0
    
    gm_ds_qs = _filter_date(DirectSalePayment.objects.filter(status='confirmed'))
    gm_revenue = float(gm_ds_qs.aggregate(t=Sum('total_sale_value'))['t'] or 0)
    gm_received = float(gm_ds_qs.aggregate(t=Sum('amount_received_cash') + Sum('amount_received_transfer'))['t'] or 0)
    gm_outstanding = max(0, gm_revenue - gm_received)
        
    company_total_revenue = float(total_sales_value) + float(brand_sales_value) + gm_revenue
    company_received_money = float(total_money_received) + float(brand_received) + gm_received
    company_outstanding_money = float(total_money_outstanding) + float(brand_outstanding) + gm_outstanding

    return render(request, 'reports/company_flow.html', {
        'current_user': user, 'show_money': show_money,
        'raw_maize_in': raw_maize_in, 'raw_wheat_in': raw_wheat_in,
        'raw_maize_out': raw_maize_out, 'raw_wheat_out': raw_wheat_out,
        'raw_maize_bal': raw_maize_bal, 'raw_wheat_bal': raw_wheat_bal,
        'clean_maize_in': clean_maize_in, 'clean_wheat_in': clean_wheat_in,
        'clean_maize_out': clean_maize_out, 'clean_wheat_out': clean_wheat_out,
        'clean_maize_bal': clean_maize_bal, 'clean_wheat_bal': clean_wheat_bal,
        'maize_bags_milled': maize_bags_milled, 'wheat_bags_milled': wheat_bags_milled,
        'maize_powder_produced': maize_powder_produced, 'wheat_powder_produced': wheat_powder_produced,
        'maize_powder_used': maize_powder_used, 'wheat_powder_used': wheat_powder_used,
        'maize_powder_bal': maize_powder_bal, 'wheat_powder_bal': wheat_powder_bal,
        'maize_sacks_produced': maize_sacks_produced, 'wheat_sacks_produced': wheat_sacks_produced,
        'total_sacks_produced': maize_sacks_produced + wheat_sacks_produced,
        'prod_officer_rows': prod_officer_rows,
        'fg_in': fg_in, 'fg_out': fg_out, 'fg_bal': fg_bal,
        'fg_maize_in': fg_maize_in, 'fg_wheat_in': fg_wheat_in,
        'fg_maize_out': fg_maize_out, 'fg_wheat_out': fg_wheat_out,
        'fg_maize_bal': fg_maize_bal, 'fg_wheat_bal': fg_wheat_bal,
        'total_brand_sacks': total_brand_sacks, 'brand_sales_value': brand_sales_value,
        'brand_received': brand_received, 'brand_outstanding': brand_outstanding,
        'sm_rows': sm_rows,
        'total_sacks_collected': total_sacks_collected,
        'total_goods_outstanding': total_goods_outstanding,
        'total_money_outstanding': total_money_outstanding,
        'total_money_received': total_money_received,
        'total_sales_value': total_sales_value,
        'gm_maize_hand': gm_maize_hand, 'gm_wheat_hand': gm_wheat_hand,
        'gm_user_maize': gm_user_maize, 'gm_user_wheat': gm_user_wheat,
        'md_user_maize': md_user_maize, 'md_user_wheat': md_user_wheat,
        'gm_maize_pieces': gm_maize_pieces, 'gm_wheat_pieces': gm_wheat_pieces,
        'gm_revenue': gm_revenue,
        'company_total_revenue': company_total_revenue,
        'company_received_money': company_received_money,
        'company_outstanding_money': company_outstanding_money,
        'active_range': range_type,
    })


@role_required('md')
def md_insights(request):
    """
    MD Dashboard: high level insights, charts, period analysis. Phase 8.
    """
    user = get_current_user(request)

    from procurement.models import RawMaterialReceipt, RawMaterialIssuance
    from clean_store.views import _get_clean_store_balance
    from finished_store.views import _fg_balance
    from sales.models import SalesManagerCollection, SalesManagerPayment, SalesPerson, SalesDistributionRecord, SalesResult
    from accounts.models import User
    from production.models import BrandSale
    import json
    from django.db.models.functions import TruncMonth
    from django.utils import timezone

    f_from = request.GET.get('date_from', '')
    f_to   = request.GET.get('date_to', '')

    now = timezone.now()
    if not f_from and not f_to:
        f_from = now.replace(day=1).date().isoformat()
        f_to   = now.date().isoformat()

    date_from_obj = datetime.date.fromisoformat(f_from) if f_from else None
    date_to_obj   = datetime.date.fromisoformat(f_to) if f_to else None

    # 1. Total Store Summary Layer (No money shown here)
    m_in = RawMaterialReceipt.objects.filter(material_type='maize').aggregate(t=Sum('num_bags'))['t'] or 0
    m_out = RawMaterialIssuance.objects.filter(material_type='maize').aggregate(t=Sum('num_bags_issued'))['t'] or 0
    dirty_maize = max(0, m_in - m_out)
    clean_maize = _get_clean_store_balance('maize')
    fg_maize    = _fg_balance('maize', '10kg')

    w_in = RawMaterialReceipt.objects.filter(material_type='wheat').aggregate(t=Sum('num_bags'))['t'] or 0
    w_out = RawMaterialIssuance.objects.filter(material_type='wheat').aggregate(t=Sum('num_bags_issued'))['t'] or 0
    dirty_wheat = max(0, w_in - w_out)
    clean_wheat = _get_clean_store_balance('wheat')
    fg_wheat    = _fg_balance('wheat', '10kg')

    # 2. Who is holding what
    sales_managers = User.objects.filter(role='sales_manager', status='active')
    sm_holdings = []
    for sm in sales_managers:
        from sales.views import get_sm_goods_holding, get_sm_money_outstanding
        goods = get_sm_goods_holding(sm)
        money = get_sm_money_outstanding(sm)
        if goods > 0 or money > 0:
            sm_holdings.append({
                'name': sm.full_name,
                'goods': goods,
                'money': money
            })

    # 3. Best/Worst Sales Persons (Performance)
    all_sps = SalesPerson.objects.filter(status='active')
    sp_ranks = []
    for sp in all_sps:
        issued_qs = SalesDistributionRecord.objects.filter(sales_person=sp)
        res_qs = SalesResult.objects.filter(sales_person=sp)
        
        if date_from_obj:
            issued_qs = issued_qs.filter(date__gte=date_from_obj)
            res_qs = res_qs.filter(date__gte=date_from_obj)
        if date_to_obj:
            issued_qs = issued_qs.filter(date__lte=date_to_obj)
            res_qs = res_qs.filter(date__lte=date_to_obj)

        issued = issued_qs.aggregate(t=Sum('qty_given'))['t'] or 0
        sold = float(sum(r.equivalent_sacks_sold for r in res_qs))
        pct = (sold / float(issued) * 100) if issued > 0 else 0
        if issued > 0:
            sp_ranks.append({
                'name': sp.name,
                'channel': sp.get_channel_display(),
                'issued': issued,
                'sold': sold,
                'pct': round(pct, 1)
            })
    sp_ranks.sort(key=lambda x: x['sold'], reverse=True)
    top_sps = sp_ranks[:5]

    # 4. Chart Data: Monthly Sales vs Collections
    # Group accepted collections by month
    monthly_colls = SalesManagerCollection.objects.filter(status='accepted').annotate(
        month=TruncMonth('date')
    ).values('month').annotate(total=Sum('qty_sacks')).order_by('month')

    monthly_sales = SalesResult.objects.all()
    # Manual monthly aggregation for property-based Sum
    # (Since property-based Sum isn't possible in ORM aggregate)
    sales_by_month = {}
    for r in monthly_sales:
        m_key = r.date.replace(day=1)
        sales_by_month[m_key] = sales_by_month.get(m_key, 0) + r.equivalent_sacks_sold

    months_list = []
    colls_list = []
    sales_list = []

    # Simple merging for charts (assuming current year focus)
    this_year = datetime.date.today().year
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    
    # Initialize 12 months with 0
    chart_data = {m: {'colls': 0, 'sales': 0} for m in month_names}

    for c in monthly_colls:
        if c['month'] and c['month'].year == this_year:
            m_name = c['month'].strftime('%b')
            if m_name in chart_data:
                chart_data[m_name]['colls'] = c['total']
                
    for m_key, val in sales_by_month.items():
        if m_key.year == this_year:
            m_name = m_key.strftime('%b')
            if m_name in chart_data:
                chart_data[m_name]['sales'] = float(val)

    for m in month_names:
        months_list.append(m)
        colls_list.append(chart_data[m]['colls'])
        sales_list.append(chart_data[m]['sales'])

    # 4b. Chart Data: Weekly Sales vs Collections (Current Month)
    today = datetime.date.today()
    this_month = today.month
    
    current_month_colls = SalesManagerCollection.objects.filter(status='accepted', date__year=this_year, date__month=this_month)
    current_month_sales = SalesResult.objects.filter(date__year=this_year, date__month=this_month)
    
    weeks_list = ['Week 1', 'Week 2', 'Week 3', 'Week 4', 'Week 5']
    weekly_chart_data = {w: {'colls': 0, 'sales': 0} for w in weeks_list}
    
    for c in current_month_colls:
        if c.date:
            w_idx = (c.date.day - 1) // 7
            w_name = f'Week {w_idx + 1}'
            if w_name in weekly_chart_data:
                weekly_chart_data[w_name]['colls'] += c.qty_sacks
            
    for s in current_month_sales:
        if s.date:
            w_idx = (s.date.day - 1) // 7
            w_name = f'Week {w_idx + 1}'
            if w_name in weekly_chart_data:
                weekly_chart_data[w_name]['sales'] += float(s.equivalent_sacks_sold)
                
    w_colls_list = [weekly_chart_data[w]['colls'] for w in weeks_list]
    w_sales_list = [weekly_chart_data[w]['sales'] for w in weeks_list]

    # 5. Brand Recovery (Filtered)
    brand_qs = BrandSale.objects.all()
    if date_from_obj: brand_qs = brand_qs.filter(date__gte=date_from_obj)
    if date_to_obj:   brand_qs = brand_qs.filter(date__lte=date_to_obj)
    
    brand_revenue = float(brand_qs.aggregate(t=Sum('total_amount'))['t'] or 0)
    brand_received = float(brand_qs.aggregate(t=Sum('amount_cash'))['t'] or 0) + float(brand_qs.aggregate(t=Sum('amount_transfer'))['t'] or 0)
    brand_outstanding = max(0, brand_revenue - brand_received)

    # Product Sales Mix (Maize vs Wheat)
    from sales.models import SalesResult
    prod_mix_qs = SalesResult.objects.all()
    if date_from_obj: prod_mix_qs = prod_mix_qs.filter(date__gte=date_from_obj)
    if date_to_obj:   prod_mix_qs = prod_mix_qs.filter(date__lte=date_to_obj)
    
    mix_data = {'maize': 0.0, 'wheat': 0.0}
    for r in prod_mix_qs:
        mix_data[r.material_type] += float(r.equivalent_sacks_sold)
    
    product_labels = [k.title() for k in mix_data.keys()]
    product_data = list(mix_data.values())


    context = {
        'current_user': user,
        'dirty_maize': dirty_maize,
        'clean_maize': clean_maize,
        'fg_maize': fg_maize,
        'dirty_wheat': dirty_wheat,
        'clean_wheat': clean_wheat,
        'fg_wheat': fg_wheat,
        'sm_holdings': sm_holdings,
        'top_sps': top_sps,
        'months_json': json.dumps(months_list),
        'colls_json': json.dumps(colls_list),
        'sales_json': json.dumps(sales_list),
        'weeks_json': json.dumps(weeks_list),
        'w_colls_json': json.dumps(w_colls_list),
        'w_sales_json': json.dumps(w_sales_list),
        'brand_revenue': brand_revenue,
        'brand_received': brand_received,
        'brand_outstanding': brand_outstanding,
        'product_labels_json': json.dumps(product_labels),
        'product_data_json': json.dumps(product_data),
        'this_year': this_year,
        'this_month_name': today.strftime('%B'),
        'f_from': f_from, 'f_to': f_to,
        'active_range': range_type,
    }

    return render(request, 'reports/md_insights.html', context)


@role_required('md')
def financial_summary(request):
    """
    The Financial Intelligence Center (MD only).
    Calculates P&L, SM Accountability, and Trend data.
    """
    user = get_current_user(request)
    from django.db.models import Sum, Q
    from django.utils import timezone
    from procurement.models import RawMaterialReceipt, MATERIAL_CHOICES
    from production.models import PackagingBatch
    from finished_store.views import _fg_balance
    from sales.models import SalesManagerCollection, SalesManagerPayment, DirectSalePayment, SalesResult
    from pricing.models import PackagingCostConfig, OperationalExpense, PriceConfig
    from accounts.models import User
    import json

    # --- Intelligence Window logic ---
    from datetime import timedelta
    range_type = request.GET.get('range', 'month')
    now = timezone.now()
    
    start_date = None
    end_date = now.date()

    if range_type == 'week':
        start_date = (now - timedelta(days=now.weekday())).date()
    elif range_type == 'year':
        start_date = now.replace(month=1, day=1).date()
    elif range_type == 'all':
        start_date = None
        end_date = None
    else: # Default: month
        start_date = now.replace(day=1).date()

    # Manual Overrides
    f_from = request.GET.get('date_from', '')
    f_to   = request.GET.get('date_to', '')

    if f_from: 
        try: start_date = datetime.date.fromisoformat(f_from)
        except: pass
    if f_to: 
        try: end_date = datetime.date.fromisoformat(f_to)
        except: pass

    # Synchronize string versions for template compatibility
    f_from = start_date.isoformat() if start_date else ''
    f_to = end_date.isoformat() if end_date else ''

    date_from_obj = start_date
    date_to_obj   = end_date

    # 1. High Level Material-Specific Metrics
    materials = ['maize', 'wheat']
    metrics = {mat: {
        'produced': 0, 'sold': 0, 'in_store': 0, 'sm_holding': 0,
        'revenue': 0, 'outstanding': 0
    } for mat in materials}

    for mat in materials:
        # Produced (Filtered)
        prod_qs = PackagingBatch.objects.filter(material_type=mat)
        if date_from_obj: prod_qs = prod_qs.filter(date__gte=date_from_obj)
        if date_to_obj:   prod_qs = prod_qs.filter(date__lte=date_to_obj)
        metrics[mat]['produced'] = prod_qs.aggregate(t=Sum('qty_10kg'))['t'] or 0
        
        # In Store (Physical balance - always all-time)
        metrics[mat]['in_store'] = _fg_balance(mat, '10kg')

        # Sold (Filtered)
        sm_res_qs = SalesResult.objects.filter(material_type=mat)
        if date_from_obj: sm_res_qs = sm_res_qs.filter(date__gte=date_from_obj)
        if date_to_obj:   sm_res_qs = sm_res_qs.filter(date__lte=date_to_obj)
        sm_sold_eq = float(sum(r.equivalent_sacks_sold for r in sm_res_qs))
        
        ds_sold_qs = DirectSalePayment.objects.filter(material_type=mat, status='confirmed')
        if date_from_obj: ds_sold_qs = ds_sold_qs.filter(date__gte=date_from_obj)
        if date_to_obj:   ds_sold_qs = ds_sold_qs.filter(date__lte=date_to_obj)
        ds_sold = ds_sold_qs.aggregate(t=Sum('qty_sold'))['t'] or 0
        metrics[mat]['sold'] = sm_sold_eq + float(ds_sold)

        # Outstanding with SMs (Goods in their hands)
        from sales.views import get_sm_goods_holding
        total_sm_holding = 0
        for sm in User.objects.filter(role='sales_manager', status='active'):
            total_sm_holding += get_sm_goods_holding(sm, mat)
        metrics[mat]['sm_holding'] = total_sm_holding

        # Revenue (Filtered)
        sm_rev_qs = SalesResult.objects.filter(material_type=mat)
        if date_from_obj: sm_rev_qs = sm_rev_qs.filter(date__gte=date_from_obj)
        if date_to_obj:   sm_rev_qs = sm_rev_qs.filter(date__lte=date_to_obj)
        sm_rev = sm_rev_qs.aggregate(t=Sum('gross_value'))['t'] or 0

        ds_rev_qs = DirectSalePayment.objects.filter(material_type=mat, status='confirmed')
        if date_from_obj: ds_rev_qs = ds_rev_qs.filter(date__gte=date_from_obj)
        if date_to_obj:   ds_rev_qs = ds_rev_qs.filter(date__lte=date_to_obj)
        ds_rev = ds_rev_qs.aggregate(t=Sum('total_sale_value'))['t'] or 0
        
        metrics[mat]['revenue'] = float(sm_rev) + float(ds_rev)

        # Outstanding Money
        from sales.views import get_sm_money_outstanding
        total_sm_money = 0
        for sm in User.objects.filter(role='sales_manager', status='active'):
            total_sm_money += get_sm_money_outstanding(sm) # Note: SM money is mixed, we estimate proportional link in the accountability table
        
        ds_out = DirectSalePayment.objects.filter(material_type=mat, status='confirmed')
        total_ds_out = sum(float(ds.outstanding) for ds in ds_out)
        
        # For the global dashboard cards, we just sum them
        # Note: SM outstanding is better handled in the table because it's hard to split perfectly by material 
        # unless we use the ratio logic from the accountability table.
        metrics[mat]['outstanding_ds'] = total_ds_out

    # 2. Sales Accountability Table (Per SM)
    sm_rows = []
    for sm in User.objects.filter(role='sales_manager', status='active').order_by('full_name'):
        for mat in materials:
            # Sacks Collected (Total ever)
            coll_qs = SalesManagerCollection.objects.filter(sales_manager=sm, material_type=mat, status='accepted')
            collected = coll_qs.aggregate(t=Sum('qty_sacks'))['t'] or 0
            coll_val = coll_qs.aggregate(t=Sum('total_value'))['t'] or 0
            
            # Sacks Sold (Results recorded - converted to equivalent sacks)
            res_qs = SalesResult.objects.filter(recorded_by=sm, material_type=mat)
            sold = float(sum(r.equivalent_sacks_sold for r in res_qs))
            comm = res_qs.aggregate(t=Sum('commission_amount'))['t'] or 0
            
            # Outstanding Sacks
            holding = get_sm_goods_holding(sm, mat)
            
            # Financials (Proportional split for SM payments)
            # This logic matches the dashboard.html logic for consistency
            m_net_owed = float(SalesManagerCollection.objects.filter(sales_manager=sm, material_type='maize', status='accepted').aggregate(t=Sum('total_value'))['t'] or 0) - float(SalesResult.objects.filter(recorded_by=sm, material_type='maize').aggregate(t=Sum('commission_amount'))['t'] or 0)
            w_net_owed = float(SalesManagerCollection.objects.filter(sales_manager=sm, material_type='wheat', status='accepted').aggregate(t=Sum('total_value'))['t'] or 0) - float(SalesResult.objects.filter(recorded_by=sm, material_type='wheat').aggregate(t=Sum('commission_amount'))['t'] or 0)
            total_net = m_net_owed + w_net_owed
            
            total_paid = float(SalesManagerPayment.objects.filter(sales_manager=sm, status='confirmed').aggregate(t=Sum('amount_cash') + Sum('amount_transfer'))['t'] or 0)
            
            if total_net > 0:
                ratio = (m_net_owed if mat == 'maize' else w_net_owed) / total_net
            else:
                ratio = 0.5
            
            mat_paid = total_paid * ratio
            mat_owed = (m_net_owed if mat == 'maize' else w_net_owed) - mat_paid
            
            if collected > 0 or holding > 0:
                sm_rows.append({
                    'sm': sm,
                    'material': mat,
                    'collected': collected,
                    'sold': sold,
                    'holding': holding,
                    'paid': mat_paid,
                    'outstanding': max(0, mat_owed)
                })

    # 3. GM DIRECT SALES ACCOUNTABILITY
    from sales.models import GMRemittance, DirectSalePayment
    gm_managers = User.objects.filter(role='manager', status='active')
    gm_rows = []
    total_gm_received = 0
    total_gm_remitted = 0

    for gm in gm_managers:
        gs_received = DirectSalePayment.objects.filter(
            recorded_by=gm, status__in=['pending_md', 'confirmed']
        ).aggregate(t=Sum('amount_received_cash') + Sum('amount_received_transfer'))['t'] or 0
        gs_remitted = GMRemittance.objects.filter(
            recorded_by=gm, status='confirmed'
        ).aggregate(t=Sum('amount_cash') + Sum('amount_transfer'))['t'] or 0
        pending_remittance = GMRemittance.objects.filter(
            recorded_by=gm, status='pending_md'
        ).aggregate(t=Sum('amount_cash') + Sum('amount_transfer'))['t'] or 0

        gm_rows.append({
            'gm': gm,
            'received': float(gs_received),
            'remitted': float(gs_remitted),
            'pending': float(pending_remittance),
            'outstanding': float(gs_received) - float(gs_remitted)
        })
        total_gm_received += float(gs_received)
        total_gm_remitted += float(gs_remitted)

    # 4. Profit & Loss Engine
    # Revenue
    total_rev = sum(m['revenue'] for m in metrics.values())
    
    # Costs
    # a. Raw Material Cost (MD-approved costs for all RECEIPTS in the period)
    rm_qs = RawMaterialReceipt.objects.filter(cost_status='approved')
    if date_from_obj: rm_qs = rm_qs.filter(date__gte=date_from_obj)
    if date_to_obj:   rm_qs = rm_qs.filter(date__lte=date_to_obj)
    raw_mat_cost = float(rm_qs.aggregate(t=Sum('total_cost'))['t'] or 0)
    
    # b. Packaging Cost (Produced Sacks * Active Packaging Cost at production time)
    # Since accurate per-batch matching is complex, we use total produced * current active packaging cost as a strong estimate
    packaging_cost = 0
    active_pack_config = PackagingCostConfig.get_active_config(timezone.now().date())
    unit_pack_cost = 0
    if active_pack_config:
        unit_pack_cost = float(active_pack_config.cost_per_sack) + float(active_pack_config.nylon_cost_per_piece)
        
    for mat in materials:
        packaging_cost += metrics[mat]['produced'] * unit_pack_cost
    
    # c. Operational Expenses (Filtered)
    oe_qs = OperationalExpense.objects.all()
    if date_from_obj: oe_qs = oe_qs.filter(date__gte=date_from_obj)
    if date_to_obj:   oe_qs = oe_qs.filter(date__lte=date_to_obj)
    op_expenses = float(oe_qs.aggregate(t=Sum('amount'))['t'] or 0)
    
    total_costs = raw_mat_cost + packaging_cost + op_expenses
    net_profit = total_rev - total_costs
    
    profit_margin_pct = (net_profit / total_rev * 100) if total_rev > 0 else 0
    profit_ratio = (net_profit / total_rev) if total_rev > 0 else 0

    # 4. Trend Charts (Dynamic period aggregation)
    # 4. Trend Charts (Dynamic period aggregation)
    labels = []
    prod_data = []
    sales_data = []
    rev_data = []
    profit_data = []
    
    # Granularity based on Intelligence Window
    chart_period = range_type
    if chart_period == 'year':
        ranges = 3
    elif chart_period == 'month':
        ranges = 6
    else: # week or all
        chart_period = 'week'
        ranges = 5

    today = timezone.now().date()
    
    for i in range(ranges - 1, -1, -1):
        if chart_period == 'year':
            start_date = datetime.date(today.year - i, 1, 1)
            end_date = datetime.date(today.year - i, 12, 31)
            label = str(start_date.year)
        elif chart_period == 'month':
            target_month = (today.month - i - 1) % 12 + 1
            target_year = today.year + (today.month - i - 1) // 12
            start_date = datetime.date(target_year, target_month, 1)
            if target_month == 12:
                end_date = datetime.date(target_year, 12, 31)
            else:
                end_date = datetime.date(target_year, target_month + 1, 1) - timezone.timedelta(days=1)
            label = start_date.strftime('%b %Y')
        else:
            target_date = today - timezone.timedelta(weeks=i)
            start_date = target_date - timezone.timedelta(days=target_date.weekday())
            end_date = start_date + timezone.timedelta(days=6)
            label = f"Wk {start_date.strftime('%d %b')}"
        
        labels.append(label)
        
        # Production
        w_prod = PackagingBatch.objects.filter(date__range=[start_date, end_date]).aggregate(t=Sum('qty_10kg'))['t'] or 0
        prod_data.append(int(w_prod))
        
        # Sales
        w_res = SalesResult.objects.filter(date__range=[start_date, end_date])
        w_sm_sold = float(sum(r.equivalent_sacks_sold for r in w_res))
        w_ds_sold = DirectSalePayment.objects.filter(date__range=[start_date, end_date], status='confirmed').aggregate(t=Sum('qty_sold'))['t'] or 0
        sales_data.append(int(w_sm_sold + w_ds_sold))
        
        # Revenue
        w_sm_rev = SalesResult.objects.filter(date__range=[start_date, end_date]).aggregate(t=Sum('gross_value'))['t'] or 0
        w_ds_rev = DirectSalePayment.objects.filter(date__range=[start_date, end_date], status='confirmed').aggregate(t=Sum('total_sale_value'))['t'] or 0
        w_rev = float(w_sm_rev) + float(w_ds_rev)
        rev_data.append(w_rev)
        
        # Expenses
        w_exp = OperationalExpense.objects.filter(date__range=[start_date, end_date]).aggregate(t=Sum('amount'))['t'] or 0
        profit_data.append(w_rev - float(w_exp))

    chart_data = {
        'labels': labels,
        'production': prod_data,
        'sales': sales_data,
        'revenue': rev_data,
        'profit': profit_data,
    }

    # Filter the lists shown at the bottom too
    recent_expenses = OperationalExpense.objects.all()
    pending_raw = RawMaterialReceipt.objects.filter(cost_status='pending')
    if date_from_obj:
        recent_expenses = recent_expenses.filter(date__gte=date_from_obj)
        pending_raw = pending_raw.filter(date__gte=date_from_obj)
    if date_to_obj:
        recent_expenses = recent_expenses.filter(date__lte=date_to_obj)
        pending_raw = pending_raw.filter(date__lte=date_to_obj)

    return render(request, 'reports/financial_summary.html', {
        'current_user': user,
        'f_from': f_from, 'f_to': f_to,
        'active_range': range_type,
        'chart_data_json': json.dumps(chart_data),
        'metrics': metrics,
        'sm_rows': sm_rows,
        'gm_rows': gm_rows,
        'total_gm_received': total_gm_received,
        'total_gm_remitted': total_gm_remitted,
        'total_rev': total_rev,
        'raw_mat_cost': raw_mat_cost,
        'packaging_cost': packaging_cost,
        'op_expenses': op_expenses,
        'total_costs': total_costs,
        'net_profit': net_profit,
        'profit_margin_pct': profit_margin_pct,
        'profit_ratio_pct': profit_ratio * 100,
        'chart_data_json': json.dumps(chart_data),
        'expenses_list': recent_expenses.order_by('-date')[:10],
        'pending_raw_costs': pending_raw.order_by('-date'),
    })


@role_required('manager', 'md')
def md_ledger(request):
    """
    Dedicated ledger view for the MD with hierarchical drill-down:
    Roles -> Staff Members -> Individual Records.
    Excludes technical logins/logouts.
    """
    user = get_current_user(request)
    
    # 1. Drill-down parameters
    f_role = request.GET.get('role_filter', '')
    f_user = request.GET.get('user_id', '')
    
    # 2. Standardized Intelligence Window
    from django.utils import timezone
    from datetime import timedelta
    range_type = request.GET.get('range', 'month')
    now = timezone.now()
    
    start_date = None
    end_date = now.date()

    if range_type == 'week':
        start_date = (now - timedelta(days=now.weekday())).date()
    elif range_type == 'year':
        start_date = now.replace(month=1, day=1).date()
    elif range_type == 'all':
        start_date = None
        end_date = None
    else: # Default: month
        start_date = now.replace(day=1).date()

    # Manual Overrides
    f_from = request.GET.get('date_from', '')
    f_to = request.GET.get('date_to', '')
    if f_from: 
        try: start_date = datetime.date.fromisoformat(f_from)
        except: pass
    if f_to: 
        try: end_date = datetime.date.fromisoformat(f_to)
        except: pass
    
    f_search = request.GET.get('search', '').strip()

    excluded_actions = ['LOGIN', 'LOGOUT', 'IMPERSONATE_START', 'IMPERSONATE_STOP']
    
    # Establish base logs with role-based restriction
    base_logs = AuditLog.objects.exclude(action__in=excluded_actions)
    if user.role == 'manager':
        base_logs = base_logs.exclude(user_role='md')
    
    # Pre-fetch roles
    ROLE_MAP = {
        'store_officer': 'Store Officer',
        'production_officer': 'Production Officer',
        'sales_manager': 'Sales Manager',
        'manager': 'General Manager',
        'md': 'Managing Director',
    }
    
    context = {
        'current_user': user,
        'f_role': f_role,
        'f_user': f_user,
        'f_from': f_from,
        'f_to': f_to,
        'f_search': f_search,
        'role_map': ROLE_MAP,
        'active_range': range_type,
    }

    if not f_role and not f_user:
        # LEVEL 1: Select a Role
        raw_roles = base_logs.values_list('user_role', flat=True).distinct().order_by('user_role')
        context['available_roles'] = [(r, ROLE_MAP.get(r, r.replace('_', ' ').title())) for r in raw_roles if r]
        context['view_level'] = 'roles'
        
    elif f_role and not f_user:
        # LEVEL 2: Select a Staff Member in that role
        context['staff_members'] = base_logs.filter(user_role=f_role).values('user_id', 'user_name').distinct().order_by('user_name')
        context['role_name'] = ROLE_MAP.get(f_role, f_role.replace('_', ' ').title())
        context['view_level'] = 'staff'
        
    else:
        # LEVEL 3: View Records for a specific user
        logs = base_logs.filter(user_id=f_user).order_by('-timestamp')
        
        # Apply Intelligence Window / Manual Filters
        if start_date:
            logs = logs.filter(timestamp__date__gte=start_date)
        if end_date:
            logs = logs.filter(timestamp__date__lte=end_date)
        if f_search:
            from django.db.models import Q
            logs = logs.filter(Q(description__icontains=f_search) | Q(action__icontains=f_search))

        # Add friendly names and enrich with current status
        staff_name = "Unknown Staff"
        staff_role = "Unknown Role"
        
        # Pre-fetch statuses to avoid N+1 queries for the same object
        # We'll just do simple fetching for now as the ledger is usually filtered to one person anyway.
        for l in logs:
            l.friendly_action = l.action.replace('_', ' ').title()
            staff_name = l.user_name
            staff_role = ROLE_MAP.get(l.user_role, l.user_role.replace('_', ' ').title())
            
            # Enrich with real-time status if it's a "handshake" object
            if l.object_type and l.object_id:
                try:
                    obj = None
                    if l.object_type == 'SalesManagerCollection':
                        from sales.models import SalesManagerCollection
                        obj = SalesManagerCollection.objects.filter(pk=l.object_id).first()
                    elif l.object_type == 'FinishedGoodsReceipt':
                        from finished_store.models import FinishedGoodsReceipt
                        obj = FinishedGoodsReceipt.objects.filter(pk=l.object_id).first()
                    elif l.object_type == 'SalesManagerPayment':
                        from sales.models import SalesManagerPayment
                        obj = SalesManagerPayment.objects.filter(pk=l.object_id).first()
                    elif l.object_type == 'DirectSalePayment':
                        from sales.models import DirectSalePayment
                        obj = DirectSalePayment.objects.filter(pk=l.object_id).first()
                    
                    if obj and hasattr(obj, 'status'):
                        l.current_status = obj.status.replace('_', ' ').title()
                except:
                    pass

        context['logs'] = logs
        context['staff_name'] = staff_name
        context['staff_role'] = staff_role
        context['view_level'] = 'records'

    return render(request, 'reports/md_ledger.html', context)


@role_required('manager', 'md')
def record_monthly_snapshot(request):
    """
    MD/Manager-triggered closing of the month.
    Aggregates all company balances and saves a permanent snapshot.
    """
    user = get_current_user(request)
    now = datetime.datetime.now()
    
    # Default to previous month if today is early in the month, or current month
    if now.day <= 5:
        prev = now.replace(day=1) - datetime.timedelta(days=1)
        default_year, default_month = prev.year, prev.month
    else:
        default_year, default_month = now.year, now.month

    # Get from POST or use defaults
    try:
        year = int(request.POST.get('year', default_year))
        month = int(request.POST.get('month', default_month))
    except (ValueError, TypeError):
        year, month = default_year, default_month

    # --- AGGREGATE DATA ---
    from clean_store.views import _get_clean_store_balance
    from finished_store.views import _fg_balance
    from sales.views import get_sm_goods_holding, get_sm_money_outstanding, get_gm_goods_holding, get_company_goods_holding
    from production.views import get_sacks_in_hand, get_sacks_in_transit
    from sales.models import CompanyRetailLedger, DirectSalePayment
    from procurement.models import RawMaterialReceipt, RawMaterialIssuance

    # 1. Store
    raw_m_in = RawMaterialReceipt.objects.filter(material_type='maize').aggregate(t=Sum('num_bags'))['t'] or 0
    raw_m_out = RawMaterialIssuance.objects.filter(material_type='maize').aggregate(t=Sum('num_bags_issued'))['t'] or 0
    dirty_maize = max(0, raw_m_in - raw_m_out)

    raw_w_in = RawMaterialReceipt.objects.filter(material_type='wheat').aggregate(t=Sum('num_bags'))['t'] or 0
    raw_w_out = RawMaterialIssuance.objects.filter(material_type='wheat').aggregate(t=Sum('num_bags_issued'))['t'] or 0
    dirty_wheat = max(0, raw_w_in - raw_w_out)

    clean_maize = _get_clean_store_balance('maize')
    clean_wheat = _get_clean_store_balance('wheat')

    # 2. Production
    total_maize_hand = 0
    total_wheat_hand = 0
    total_maize_transit = 0
    total_wheat_transit = 0
    
    prod_officers = User.objects.filter(role='production_officer', status='active')
    for po in prod_officers:
        total_maize_hand += get_sacks_in_hand(po, 'maize')
        total_wheat_hand += get_sacks_in_hand(po, 'wheat')
        total_maize_transit += get_sacks_in_transit(po, 'maize')
        total_wheat_transit += get_sacks_in_transit(po, 'wheat')

    # Monthly production activity
    prod_m_milled = MillingBatch.objects.filter(material_type='maize', date__year=year, date__month=month).aggregate(
        t=Sum('bags_milled_new') + Sum('outstanding_bags_milled')
    )['t'] or 0
    prod_w_milled = MillingBatch.objects.filter(material_type='wheat', date__year=year, date__month=month).aggregate(
        t=Sum('bags_milled_new') + Sum('outstanding_bags_milled')
    )['t'] or 0
    prod_m_pack = PackagingBatch.objects.filter(material_type='maize', date__year=year, date__month=month).aggregate(t=Sum('qty_10kg'))['t'] or 0
    prod_w_pack = PackagingBatch.objects.filter(material_type='wheat', date__year=year, date__month=month).aggregate(t=Sum('qty_10kg'))['t'] or 0

    # 3. FG Store
    fg_m_10kg = _fg_balance('maize', '10kg')
    fg_w_10kg = _fg_balance('wheat', '10kg')

    # 4. Sales Manager
    sm_m_holding = 0
    sm_w_holding = 0
    sm_money_out = 0
    for sm in User.objects.filter(role='sales_manager', status='active'):
        sm_m_holding += get_sm_goods_holding(sm, 'maize')
        sm_w_holding += get_sm_goods_holding(sm, 'wheat')
        sm_money_out += float(get_sm_money_outstanding(sm))

    # 5. GM
    gm_m_hand = get_gm_goods_holding('maize')
    gm_w_hand = get_gm_goods_holding('wheat')
    gm_ret_m = CompanyRetailLedger.objects.filter(material_type='maize').aggregate(t=Sum('pieces_changed'))['t'] or 0
    gm_ret_w = CompanyRetailLedger.objects.filter(material_type='wheat').aggregate(t=Sum('pieces_changed'))['t'] or 0
    
    ds_out = 0
    ds_qs = DirectSalePayment.objects.filter(status='confirmed')
    for ds in ds_qs:
        ds_out += float(ds.outstanding)

    # --- SAVE SNAPSHOT ---
    with transaction.atomic():
        snapshot, created = MonthlySnapshot.objects.update_or_create(
            year=year, month=month,
            defaults={
                'clean_maize_bags': int(clean_maize),
                'clean_wheat_bags': int(clean_wheat),
                'dirty_maize_bags': int(dirty_maize),
                'dirty_wheat_bags': int(dirty_wheat),
                'prod_maize_hand': int(total_maize_hand),
                'prod_wheat_hand': int(total_wheat_hand),
                'prod_maize_transit': int(total_maize_transit),
                'prod_wheat_transit': int(total_wheat_transit),
                'prod_maize_milled_bags': int(prod_m_milled),
                'prod_wheat_milled_bags': int(prod_w_milled),
                'prod_maize_packaged_sacks': int(prod_m_pack),
                'prod_wheat_packaged_sacks': int(prod_w_pack),
                'fg_maize_10kg': int(fg_m_10kg),
                'fg_wheat_10kg': int(fg_w_10kg),
                'sm_maize_holding': int(sm_m_holding),
                'sm_wheat_holding': int(sm_w_holding),
                'sm_money_outstanding': sm_money_out,
                'gm_maize_hand': int(gm_m_hand),
                'gm_wheat_hand': int(gm_w_hand),
                'gm_retail_pieces_maize': int(gm_ret_m),
                'gm_retail_pieces_wheat': int(gm_ret_w),
                'gm_direct_sale_outstanding': ds_out,
                'recorded_by': user,
            }
        )

        # --- LOG TO LEDGER ---
        import calendar
        month_name = calendar.month_name[month]
        # GM Liability
        from sales.models import GMRemittance
        gm_received_all = DirectSalePayment.objects.filter(status__in=['pending_md', 'confirmed']).aggregate(t=Sum('amount_received_cash') + Sum('amount_received_transfer'))['t'] or 0
        gm_remitted_all = GMRemittance.objects.filter(status='confirmed').aggregate(t=Sum('amount_cash') + Sum('amount_transfer'))['t'] or 0
        gm_pending_all = GMRemittance.objects.filter(status='pending_md').aggregate(t=Sum('amount_cash') + Sum('amount_transfer'))['t'] or 0
        gm_debt = float(gm_received_all) - float(gm_remitted_all)

        description = (
            f"COMPANY-WIDE MONTH-END CLOSING RECORD: {month_name} {year}. | "
            f"PRODUCTION: Maize {total_maize_hand} Hand/{total_maize_transit} Transit, Wheat {total_wheat_hand} Hand/{total_wheat_transit} Transit. | "
            f"STORE: Maize {clean_maize} Clean/{dirty_maize} Dirty, Wheat {clean_wheat} Clean/{dirty_wheat} Dirty. | "
            f"SALES: Sacks Holding {sm_m_holding + sm_w_holding}, Money Owed ₦{sm_money_out:,.2f}. | "
            f"GM: Sacks {gm_m_hand + gm_w_hand}, Direct Sales Case: Received ₦{gm_received_all:,.2f}, Remitted ₦{gm_remitted_all:,.2f}, Pending ₦{gm_pending_all:,.2f}, Owed ₦{gm_debt:,.2f}."
        )
        from audit.utils import log_action
        log_action(request, user, 'REPORTS', 'MONTHLY_SNAPSHOT', description, 'MonthlySnapshot', snapshot.pk)

    messages.success(request, f"Company-wide monthly snapshot for {month_name} {year} has been successfully recorded into the ledger.")
    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/reports/dashboard/'))
